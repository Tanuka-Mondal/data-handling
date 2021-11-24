import speech_recognition as sr
r = sr.Recognizer()

with sr.Microphone() as source:
    r.pause_threshold = 1
    r.adjust_for_ambient_noise(source)
    print("Say Something")
    audio_text = r.listen(source)
    print("Time over, thanks")

    a = r.recognize_google(audio_text)
   
    try:
        print("Text: "+r.recognize_google(audio_text))
    except:
         print("Sorry, I did not get that")

import openpyxl

wb = openpyxl.load_workbook("weather.xlsx")
ws = wb.active
i = 0
v = '' 

for r in range(1,ws.max_row+1):
    for c in range(1,ws.max_column+1):
        s = ws.cell(r,c).value    
        if str(s)==a:
           print('Yes')  
           print('row: '+str(r)+' and col: '+str(c))
           v = 'F'+str(r)
           
ws[v] = 'YES'
wb.save('weather.xlsx')
