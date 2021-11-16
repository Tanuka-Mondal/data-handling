import openpyxl
path = "D:\\clg_project\\weather.xlsx"
workbook = openpyxl.load_workbook(path) #read
sheet = workbook.active #store
rowNum = sheet.max_row #Number of row
print("Number of row: "+ str(rowNum)) 
colNum = sheet.max_column #Number of column
print("Number of column: "+str(colNum))
r,c = map(int,input("Enter row, column of the cell: ").split())
cellObject = sheet.cell(row = r, column = c) #one cell
print(cellObject.value)
rowVal = int(input("Enter the row you want to see: "))
for i in range(1, colNum + 1):
    rowObjects = sheet.cell(row = rowVal, column = i)
    print(rowObjects.value, end = "; ")
colVal = int(input("\nEnter the column you want to see: "))
for i in range(1, rowNum + 1):
    colObjects = sheet.cell(row = i, column = colVal)
    print(colObjects.value)
